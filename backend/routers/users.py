from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import csv
import io
import openpyxl

from database import get_db
from models import User
from schemas import UserCreate, UserUpdate, UserOut, ImportResult, RUOLI_VALIDI, _validate_ruoli
from auth import hash_password, require_admin, get_current_user
from audit import log_password_event

router = APIRouter(prefix="/api/users", tags=["users"])

# I path statici (/import/template, /import) devono stare prima di /{user_id}
# per evitare che FastAPI li interpreti come interi (HTTP 422).


@router.get("/import/template")
def import_template(_: User = Depends(require_admin)):
    righe = [
        ["nome", "cognome", "email", "azienda", "ruolo", "password"],
        ["Mario", "Rossi", "mario.rossi@example.com", "Acme Srl", "opts", "password123"],
        ["Giulia", "Bianchi", "giulia.bianchi@example.com", "Acme Srl", "validatore", "password456"],
    ]
    buf = io.StringIO()
    csv.writer(buf).writerows(righe)
    content = "﻿" + buf.getvalue()  # UTF-8 BOM per compatibilità Excel
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=template_utenti.csv"},
    )


@router.post("/import", response_model=ImportResult)
def import_users(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    filename = (file.filename or "").lower()
    creati, saltati = 0, 0
    errori: List[str] = []
    rows: List[dict] = []
    raw = file.file.read()

    if filename.endswith(".csv"):
        rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    elif filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        intestazioni = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(intestazioni, row)))
        wb.close()
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa CSV o XLSX.")

    for i, row in enumerate(rows, start=2):
        nome     = str(row.get("nome")     or "").strip()
        cognome  = str(row.get("cognome")  or "").strip()
        email    = str(row.get("email")    or "").strip().lower()
        azienda  = str(row.get("azienda")  or "").strip()
        ruolo    = str(row.get("ruolo")    or "opts").strip().lower()
        password = str(row.get("password") or "").strip()

        if not nome or not cognome or not email:
            errori.append(f"Riga {i}: nome, cognome ed email sono obbligatori")
            continue
        if ruolo not in RUOLI_VALIDI:
            errori.append(f"Riga {i}: ruolo '{ruolo}' non valido — usa {', '.join(sorted(RUOLI_VALIDI))}")
            continue
        if not password:
            errori.append(f"Riga {i}: password obbligatoria")
            continue
        if db.query(User).filter(User.email == email).first():
            saltati += 1
            continue

        new_user = User(
            nome=nome, cognome=cognome, email=email,
            password_hash=hash_password(password), azienda=azienda, ruolo=ruolo,
        )
        db.add(new_user)
        db.flush()
        log_password_event(
            action="import", actor_id=admin.id, actor_email=admin.email,
            target_id=new_user.id, target_email=email, request=request,
        )
        creati += 1

    db.commit()
    return {"creati": creati, "saltati": saltati, "errori": errori}


@router.get("/", response_model=List[UserOut])
def list_users(db: Session = Depends(get_db), me: User = Depends(get_current_user)):
    if not me.has_role("admin", "validatore", "gestore_commesse"):
        raise HTTPException(status_code=403, detail="Accesso riservato")
    return db.query(User).all()


@router.post("/", response_model=UserOut, status_code=201)
def create_user(
    payload: UserCreate, request: Request,
    db: Session = Depends(get_db), admin: User = Depends(require_admin),
):
    if db.query(User).filter(User.email == payload.email).first():
        raise HTTPException(status_code=400, detail="Email già registrata")
    ruoli = _validate_ruoli(payload.ruoli if payload.ruoli else [payload.ruolo or "opts"])
    user = User(
        nome=payload.nome, cognome=payload.cognome, email=payload.email,
        password_hash=hash_password(payload.password),
        azienda=payload.azienda or "", ditta_id=payload.ditta_id,
        ruolo=",".join(ruoli),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    log_password_event(
        action="create", actor_id=admin.id, actor_email=admin.email,
        target_id=user.id, target_email=user.email, request=request,
    )
    return user


@router.put("/{user_id}", response_model=UserOut)
def update_user(
    user_id: int, payload: UserUpdate, request: Request,
    db: Session = Depends(get_db), admin: User = Depends(require_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    data = payload.model_dump(exclude_unset=True)
    password_changed = False
    if "ruoli" in data and data["ruoli"] is not None:
        user.ruolo = ",".join(_validate_ruoli(data["ruoli"]))
    for field, value in data.items():
        if field == "ruoli":
            continue
        if field == "password" and value:
            setattr(user, "password_hash", hash_password(value))
            password_changed = True
        elif field != "password":
            setattr(user, field, value)
    db.commit()
    db.refresh(user)
    if password_changed:
        log_password_event(
            action="admin_reset", actor_id=admin.id, actor_email=admin.email,
            target_id=user.id, target_email=user.email, request=request,
        )
    return user


@router.delete("/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Non puoi eliminare il tuo account")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    db.delete(user)
    db.commit()
